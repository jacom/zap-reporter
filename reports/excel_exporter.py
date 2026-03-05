from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from scanner.models import Alert

RISK_LABELS = {3: 'High', 2: 'Medium', 1: 'Low', 0: 'Info'}
SEVERITY_FILLS = {
    3: PatternFill(start_color='FFDC3545', end_color='FFDC3545', fill_type='solid'),
    2: PatternFill(start_color='FFFD7E14', end_color='FFFD7E14', fill_type='solid'),
    1: PatternFill(start_color='FF0DCAF0', end_color='FF0DCAF0', fill_type='solid'),
    0: PatternFill(start_color='FF6C757D', end_color='FF6C757D', fill_type='solid'),
}

HEADER_FILL = PatternFill(start_color='FF142739', end_color='FF142739', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def generate_excel(scan):
    """Generate Excel report with multiple sheets."""
    wb = Workbook()
    alerts = scan.alerts.all()

    # Sheet 1: Raw Alerts
    ws1 = wb.active
    ws1.title = 'Alerts'
    headers = ['Severity', 'CVSS', 'Alert Name', 'URL', 'Parameter', 'CWE',
               'Description', 'Solution', 'Evidence']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for row_idx, alert in enumerate(alerts, 2):
        values = [
            RISK_LABELS.get(alert.risk, 'Info'),
            alert.cvss_score,
            alert.name,
            alert.url,
            alert.param,
            f'CWE-{alert.cwe_id}' if alert.cwe_id else '',
            alert.description[:500],
            alert.solution[:500],
            alert.evidence[:200],
        ]
        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col, value=val)
            cell.border = THIN_BORDER
            if col == 1:
                cell.fill = SEVERITY_FILLS.get(alert.risk, PatternFill())
                cell.font = Font(bold=True, color='FFFFFF')

    # Auto-width
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # Sheet 2: URL Summary
    ws2 = wb.create_sheet('URL Summary')
    url_data = {}
    for alert in alerts:
        base = alert.url.split('?')[0] if alert.url else 'Unknown'
        if base not in url_data:
            url_data[base] = {'high': 0, 'medium': 0, 'low': 0, 'info': 0}
        key = RISK_LABELS.get(alert.risk, 'info').lower()
        url_data[base][key] = url_data[base].get(key, 0) + 1

    headers2 = ['URL', 'High', 'Medium', 'Low', 'Info', 'Total']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_idx, (url, counts) in enumerate(url_data.items(), 2):
        total = sum(counts.values())
        values = [url, counts['high'], counts['medium'], counts['low'], counts['info'], total]
        for col, val in enumerate(values, 1):
            ws2.cell(row=row_idx, column=col, value=val).border = THIN_BORDER

    # Sheet 3: CVSS Mapping
    ws3 = wb.create_sheet('CVSS Mapping')
    headers3 = ['Alert Name', 'CWE ID', 'CVSS Score', 'CVSS Vector', 'Risk Level']
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    seen = set()
    row_idx = 2
    for alert in alerts.order_by('-cvss_score'):
        key = (alert.name, alert.cwe_id)
        if key in seen:
            continue
        seen.add(key)
        values = [alert.name, f'CWE-{alert.cwe_id}' if alert.cwe_id else '',
                  alert.cvss_score, alert.cvss_vector, RISK_LABELS.get(alert.risk, 'Info')]
        for col, val in enumerate(values, 1):
            ws3.cell(row=row_idx, column=col, value=val).border = THIN_BORDER
        row_idx += 1

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="zap_report_{scan.id}.xlsx"'
    wb.save(response)
    return response


def generate_excel_from_alerts(all_alerts, scans, target_label):
    """Generate combined Excel report from multiple scans.

    Args:
        all_alerts: Alert queryset (pre-filtered, select_related scan__target)
        scans: list of Scan objects included in report
        target_label: str — hospital/target name for filename
    """
    wb = Workbook()

    # ── Sheet 1: Scan Summary ─────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = 'Scan Summary'
    headers0 = ['Target / โรงพยาบาล', 'Tool', 'Scan Type', 'Date', 'Critical', 'High', 'Medium', 'Low', 'Info', 'Risk Score']
    for col, h in enumerate(headers0, 1):
        cell = ws0.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for row_idx, s in enumerate(scans, 2):
        values = [
            s.target.name,
            s.get_tool_display(),
            s.get_scan_type_display(),
            s.started_at.strftime('%Y-%m-%d %H:%M') if s.started_at else '',
            s.critical_count,
            s.high_count,
            s.medium_count,
            s.low_count,
            s.info_count,
            round(s.risk_score, 1),
        ]
        for col, val in enumerate(values, 1):
            cell = ws0.cell(row=row_idx, column=col, value=val)
            cell.border = THIN_BORDER

    for col in ws0.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws0.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # ── Sheet 2: All Alerts ───────────────────────────────────────────────
    ws1 = wb.create_sheet('Alerts')
    headers1 = ['Severity', 'CVSS', 'Tool', 'Alert Name', 'URL', 'Parameter',
                'CWE', 'OWASP', 'Description', 'Solution', 'Evidence']
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for row_idx, alert in enumerate(all_alerts.order_by('-risk', '-cvss_score'), 2):
        values = [
            RISK_LABELS.get(alert.risk, 'Info'),
            alert.cvss_score,
            alert.tool.upper() if alert.tool else '',
            alert.name,
            alert.url,
            alert.param,
            f'CWE-{alert.cwe_id}' if alert.cwe_id else '',
            alert.owasp_category or '',
            (alert.description or '')[:500],
            (alert.solution or '')[:500],
            (alert.evidence or '')[:200],
        ]
        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col, value=val)
            cell.border = THIN_BORDER
            if col == 1:
                cell.fill = SEVERITY_FILLS.get(alert.risk, PatternFill())
                cell.font = Font(bold=True, color='FFFFFF')

    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # ── Sheet 3: URL Summary ──────────────────────────────────────────────
    ws2 = wb.create_sheet('URL Summary')
    url_data = {}
    for alert in all_alerts:
        base = alert.url.split('?')[0] if alert.url else 'Unknown'
        if base not in url_data:
            url_data[base] = {'high': 0, 'medium': 0, 'low': 0, 'info': 0}
        key = RISK_LABELS.get(alert.risk, 'info').lower()
        url_data[base][key] = url_data[base].get(key, 0) + 1

    headers2 = ['URL', 'High', 'Medium', 'Low', 'Info', 'Total']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_idx, (url, counts) in enumerate(url_data.items(), 2):
        total = sum(counts.values())
        values = [url, counts['high'], counts['medium'], counts['low'], counts['info'], total]
        for col, val in enumerate(values, 1):
            ws2.cell(row=row_idx, column=col, value=val).border = THIN_BORDER

    # ── Sheet 4: CVSS Mapping ─────────────────────────────────────────────
    ws3 = wb.create_sheet('CVSS Mapping')
    headers3 = ['Alert Name', 'CWE ID', 'CVSS Score', 'CVSS Vector', 'Risk Level', 'OWASP']
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    seen = set()
    row_idx = 2
    for alert in all_alerts.order_by('-cvss_score'):
        key = (alert.name, alert.cwe_id)
        if key in seen:
            continue
        seen.add(key)
        values = [
            alert.name,
            f'CWE-{alert.cwe_id}' if alert.cwe_id else '',
            alert.cvss_score,
            alert.cvss_vector,
            RISK_LABELS.get(alert.risk, 'Info'),
            alert.owasp_category or '',
        ]
        for col, val in enumerate(values, 1):
            ws3.cell(row=row_idx, column=col, value=val).border = THIN_BORDER
        row_idx += 1

    safe_name = target_label[:30].replace(' ', '_').replace('/', '-')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="VA_Report_{safe_name}.xlsx"'
    wb.save(response)
    return response
